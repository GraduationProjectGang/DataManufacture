import json
import time
import pprint
from openpyxl import Workbook
from ast import literal_eval
from dataclasses import dataclass
import numpy as np
from datetime import datetime
import time
from rotate import getRotateVec

pp = pprint.PrettyPrinter(indent=4)

def scoreMatching():
    jsonPath = 'C:\\Users\\ksh04\\PythonProjects\\DataManufacture\\datacollect.json'

    with open(jsonPath, encoding= 'UTF-8') as json_file:
        usersData = json.load(json_file)
        users = usersData['user']

        wb = Workbook()
        sheet2 = wb.active
        sheet2.title = 'scoreMatching'
        
        sheet2.cell(row=1, column=1).value = "userName"
        sheet2.cell(row=1, column=2).value = "timestamp"
        sheet2.cell(row=1, column=3).value = "strCount"
        data={}
        rowNum = 2
        num = 1
        stressJsonList = []
        for userName in users:
                data[userName] = []
                for property in users[userName]:

                        if property == 'location':
                                for locID in users[userName][property]: #locationID
                                        locItem = users[userName][property][locID]
                                        latArr = [] #latitudeArray
                                        lonArr = [] #longitudeArray
                                        speedArr = [] 
                                        accuracyArr = [] 
                                        for item in locItem: #0,1,2,3
                                                if item == "locationList":
                                                        for locListItem in users[userName][property][locID]['locationList']:
                                                                latitude = locListItem['latitude']
                                                                longitude = locListItem['longitude']
                                                                speed = locListItem['speed']
                                                                accuracy = locListItem['accuracy']
                                                                latArr.append(latitude)
                                                                lonArr.append(longitude)
                                                                speedArr.append(speed)
                                                                accuracyArr.append(accuracy)
                                                        # print(f'lat: {latitude} lon:{longitude} speeD:{speed}')  
                                                if item == "timestmamp":
                                                        dateTime = users[userName][property][locID]['timestmamp']
                                                        timestamp = time.mktime(datetime.strptime(dateTime, '%Y%m%d.%H:%M:%S').timetuple())
                                                        # print(timestamp)
                                                # print(len(latArr), len(lonArr))
                                                latAverage = 0 if len(latArr) == 0 else np.mean(latArr)
                                                lonAverage = 0 if len(lonArr) == 0 else np.mean(lonArr)
                                                speedMax = 0 if len(speedArr) == 0 else np.max(speedArr)
                                                accuracyAverage = 0 if len(accuracyArr) == 0 else np.mean(accuracyArr)
                                        # print(latAverage,lonAverage)


                                        # sheet1.cell(row=rowNum, column=1).value = userName
                                        # sheet1.cell(row=rowNum, column=2).value = dateTime
                                        # sheet1.cell(row=rowNum, column=3).value = timestamp
                                        # sheet1.cell(row=rowNum, column=4).value = latAverage
                                        # sheet1.cell(row=rowNum, column=5).value = lonAverage
                                        # sheet1.cell(row=rowNum, column=6).value = speedMax
                                        # sheet1.cell(row=rowNum, column=7).value = accuracyAverage
                                        # sheet1.cell(row=rowNum, column=8).value = 1 if speedMax > 1 else 0

                                        ifMoving = 1 if speedMax > 1 else 0
                                        data[userName].append({
                                         "timestamp":timestamp,
                                         "ifMoving":ifMoving
                                        })
                        # print(data[userName])

                # if property == 'stress':
                #     for stressID in users[userName][property]: 
                #         stressItem = users[userName][property][stressID]
                #         for item in stressItem:
                #             if item == "timestamp":
                #                 timestamp = int(int(users[userName][property][stressID][item])/1000)
                #                 print(timestamp)
                #             if item == "stressCount":
                #                 stressCount = users[userName][property][stressID][item]
                #                 print(stressCount)            

                #         dict = {"userName":userName, "timestamp":timestamp, "stressCount":stressCount}
                #         json_val = json.dumps(dict)
                #         stressJsonList.append(json_val)
                #         sheet2.cell(row=rowNum, column=1).value = userName
                #         sheet2.cell(row=rowNum, column=2).value = timestamp
                #         sheet2.cell(row=rowNum, column=3).value = int(stressCount)
                #         rowNum = rowNum + 1



                            # rowNum = rowNum + 1
        rotateData = getRotateVec()
        count = 1
        for userKey in data:
                for attr in data[userKey]:  
                        for rotateUserKey in rotateData:
                                for attr_rotate in rotateData[rotateUserKey]:
                                        if userKey == rotateUserKey:
                                                if  attr['timestamp'] == attr_rotate['timestamp']:
                                                        
                                                        for iter in attr_rotate:
                                                                if iter == 'posture' or iter == 'posture_accuracy' or iter == 'std_posture' or iter == 'orientation':
                                                                        attr[iter] = attr_rotate[iter]
                                                                        
                                           
        return data



        # wb.save(filename= 'data2.xlsx')

if __name__ == "__main__":
        data = scoreMatching()
        with open('data.json', 'w') as outfile:
                json.dump(data, outfile)