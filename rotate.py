import json
from openpyxl import Workbook
import time
import math
import numpy as np
from datetime import datetime
import ast

# 1 : 눈 밑, 2 : 눈과 일직선, 3 : 눈 위

def transfer(radian):
    ret = (radian * (180 / math.pi))
    if ret >= 0:
        return ret
    else:
        return ret + 360

def posture_x(degree):
    if degree > 0 and degree <= 90:
        return 1
    elif degree > 90 and degree <= 120:
        return 2
    elif degree > 120 and degree <= 180:
        return 3
    else:
        return 0

def posture_y(degree):
    if degree >= 240 and degree <= 300:
        return 2
    elif degree >= 60 and degree <= 120:
        return 2
    elif degree <= 270:
        return 1
    elif degree >= 30:
        return 1
    elif degree >= 180 and degree <= 210:
        return 1
    else:
        return 0
        


def getRotateVec():
    jsonPath = 'C:\\Users\\sejin\\Desktop\\졸프\\datacollect.json'

    with open(jsonPath, encoding= 'UTF-8') as json_file:
        usersData = json.load(json_file)
        users = usersData['user']

        wb = Workbook()
        sheet1 = wb.active

        sheet1.title = 'data'
        sheet1.cell(row=1, column=1).value = "userKey"
        sheet1.cell(row=1, column=2).value = "posture"
        sheet1.cell(row=1, column=3).value = "posture_accuracy"
        sheet1.cell(row=1, column=4).value =  "std_posture"
        sheet1.cell(row=1, column=5).value =  "orientation"
        sheet1.cell(row=1, column=6).value = "orientation_accuracy"
        sheet1.cell(row=1, column=7).value =  "std_orientation"
        sheet1.cell(row=1, column=8).value =  "timestamp"

        # file_path = 'C:\\Users\\sejin\\Desktop\\졸프\\sample.json'
        data = {}
        rowNum = 2
        for userKey in users:

            data[userKey] = []

            for rotationVectorStress in users[userKey]:
                if rotationVectorStress == 'rotatevector':
                    for rotateVecKey in users[userKey][rotationVectorStress]:

                        data_x = np.array([])
                        data_y = np.array([])
                        std_x = 0.0
                        std_y = 0.0
                        posture_list = []
                        orientation_list = []
                        pos = 0
                        ori = 0
                        timeStamp = 0

                        x_list = []
                        y_list = []

                        flag = 0

                        for rotateVecInfo in users[userKey][rotationVectorStress][rotateVecKey]:

                            if rotateVecInfo == 'angleList':
                                for veclist in users[userKey][rotationVectorStress][rotateVecKey][rotateVecInfo]:
                                    list_data = ast.literal_eval(veclist)
                                    x_list.append(list_data[0])
                                    y_list.append(list_data[1])
                                
                                data_x = np.array(x_list)
                                data_y = np.array(y_list)

                                std_x = np.std(data_x)
                                std_y = np.std(data_y)

                                bincount_x = [0, 0, 0, 0]
                                bincount_y = [0, 0, 0]

                                for i in x_list:
                                    val = posture_x(transfer(i))
                                    posture_list.append(val)
                                    bincount_x[val] += 1
                                    
                                for i in y_list:
                                    val = posture_y(transfer(i))
                                    orientation_list.append(val)
                                    bincount_y[val] += 1

                                pos = 0
                                for i in range(0, 4):
                                    if bincount_x[i] > bincount_x[pos]:
                                        pos = i

                                # print(bincount_x)

                                ori = 0
                                for i in range(0, 3):
                                    if bincount_y[i] > bincount_y[ori]:
                                        ori = i

                                # print(bincount_y)
                                flag = 1

                            if rotateVecInfo == 'timestamp':
                                dateTime = users[userKey][rotationVectorStress][rotateVecKey][rotateVecInfo]
                                timestamp = time.mktime(datetime.strptime(dateTime, '%Y%m%d.%H:%M:%S').timetuple())
                                timeStamp = timestamp

                                flag = 2
                                # print(flag)

                            if flag == 2:
                                # print("position", pos, ", bincount_x[pos] ", bincount_x[pos], ", std_x ", std_x, ", orientation ", ori,
                                # ", bincount_y[ori] ", bincount_y[ori], ", std_y ", std_y, ", timestamp ", timeStamp)

                                data[userKey].append({
                                    "posture": int(pos),
                                    "posture_accuracy": bincount_x[pos],
                                    "std_posture": float(std_x),
                                    "orientation": int(ori),
                                    "orientation_accuracy": bincount_y[ori],
                                    "std_orientation": float(std_y),
                                    "timestamp": timeStamp
                                })

                                # sheet1.cell(row=rowNum, column=1).value = userKey
                                # sheet1.cell(row=rowNum, column=2).value = int(pos)
                                # sheet1.cell(row=rowNum, column=3).value = bincount_x[pos]
                                # sheet1.cell(row=rowNum, column=4).value = float(std_x)
                                # sheet1.cell(row=rowNum, column=5).value = int(ori)
                                # sheet1.cell(row=rowNum, column=6).value = bincount_y[ori]
                                # sheet1.cell(row=rowNum, column=7).value = float(std_y)
                                # sheet1.cell(row=rowNum, column=8).value = timeStamp

                                rowNum = rowNum + 1

                                posture_list.clear()
                                orientation_list.clear()
                                pos = 0
                                ori = 0
                                timestamp = 0
                                std_x = 0.0
                                std_y = 0.0
                                flag = 0

    return data
                        
    
    # with open(file_path, 'w') as outfile:
    #     json.dump(data, outfile)
