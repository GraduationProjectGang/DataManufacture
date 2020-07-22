import json
import copy
from openpyxl import Workbook
from collections import OrderedDict

data_json_path = 'C:\\Users\\sejin\\Documents\\GitHub\\DataManufacture\\datacollect.json'
categories_json_path = 'C:\\Users\\sejin\\Documents\\GitHub\\DataManufacture\\categories.json'

# CELL_DATE = 1
CELL_LAST_USED = 6
CELL_PACKAGE_NAME = 5 
CELL_TOTAL_TIME_IN_FOREGROUND = 4
CELL_TIMESTAMPS = 2
CELL_USER_NAME = 1
CELL_CATEGORY = 3
PADDING_STATE = 0
MAX_CATEGORIES_BY_COROUTINE = 5

categories_dict = {}

CAMERA_STRING = 'Photography'
UTILITY_ARRAY = {'Productivity', 'Beauty', 'Weather', 'News & Magazines', 'Dating', 'Tools', 'Utility'}
SNS_STRING = 'Social'
ENTERTAINMENT_ARRAY = {'Entertainment', 'Books & Reference', 'Music & Audio', 'House & Home', 'Sports', 'Video Players & Editors', 'Travel & Local', 'Lifestyle', 'Comics'}
COMMUNICATION_STRING = 'Communication'
GAME_ARRAY = {'Action', 'Racing', 'Adventure', 'Arcade', 'Puzzle', 'Simulation', 'Strategy', 'Role Playing', 'Auto & Vehicles', 'Casual', 'Card', 'Music'}
SYSTEM_STRING = 'Personalization'
EDUCATION_ARRAY = {'Education', 'Business'}
SHOPPING_STRING = 'Shopping'
MAPS_VEHICLE_ARRAY = {'Maps & Navigation', 'Auto & Vehicles'}
HEALTH_STRING = 'Health & Fitness'
FOOD_STRING = 'Food & Drink'
FINANCE_STRING = 'Finance'
BROWSER_STRING = 'Browser'

non_categorizable = set([])

#catergory의 이름(String)을 번호(Integer)로 Mapping하는 code
with open(categories_json_path, encoding= 'UTF-8') as json_file:
    categories_array = json.load(json_file)
    for app in categories_array:
        package_name = ''
        category = ''
        for app_attr in categories_array[app]:
            if app_attr == 'packageName':
                package_name = categories_array[app][app_attr]
            if app_attr == 'category':
                category = categories_array[app][app_attr]
        categories_dict[package_name] = category

#Json Parsing code
with open(data_json_path, encoding= 'UTF-8') as json_file:
    usersData = json.load(json_file)
    users = usersData['user']

    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'untitled'
    idx = 1
    jsonData = dict()
    #appending each user data
    for user_name in users:
        jsonData[user_name] = {}
        for property in users[user_name]:
            index = 0
            if property == 'usagestatsCoroutine':
                for coroutines in users[user_name][property]:
<<<<<<< HEAD
                    
=======
                    # if idx>= MAX_CATEGORIES_BY_COROUTINE:
                    #     break
>>>>>>> bf13dadfa5334b93f6d4e5c2bc1edbbad207e33b
                    sheet1.cell(idx, CELL_USER_NAME).value = user_name
                    jsonData[user_name]['userName'] = user_name
                    # print(type(jsonData[user_name]))
                    for coroutine_attr in users[user_name][property][coroutines]:
                        # if coroutine_attr == 'date':
                        #     sheet1.cell(idx, CELL_DATE).value = users[user_name][property][coroutines][coroutine_attr]
                        
                        
                        if coroutine_attr == 'timestamp':
<<<<<<< HEAD
                            sheet1.cell(idx, CELL_TIMESTAMPS).value = int((int(users[user_name][property][coroutines][coroutine_attr])/1000))
                            jsonData[user_name]['timestamp'] = int((int(users[user_name][property][coroutines][coroutine_attr])/1000))

=======
                            sheet1.cell(idx, CELL_TIMESTAMPS).value = str(int(users[user_name][property][coroutines][coroutine_attr]))
                            tempTime = str(int(users[user_name][property][coroutines][coroutine_attr]))
                            jsonData[user_name]['timestamp'] = tempTime
>>>>>>> bf13dadfa5334b93f6d4e5c2bc1edbbad207e33b
                        if coroutine_attr == 'statsList':
                            start_idx = idx
                            
                            for element in users[user_name][property][coroutines][coroutine_attr]:
                                
                                #case1 코루틴의 statsList 개수가 5개가 이상일 경우
                                uncategorizable = False
                                for stat_attr in element:
                                    # if stat_attr == 'lastTimeUsed':
                                    #     sheet1.cell(idx, CELL_LAST_USED).value = element[stat_attr]
                                    if stat_attr == 'packageName':
<<<<<<< HEAD
                                        jsonData[user_name][str(index)]={}
=======
                                        jsonData[user_name][index] = {}
>>>>>>> bf13dadfa5334b93f6d4e5c2bc1edbbad207e33b
                                        sheet1.cell(idx, CELL_PACKAGE_NAME).value = element[stat_attr]
                                        jsonData[user_name][str(index)]['packageName'] =  element[stat_attr]
                                        if element[stat_attr] in categories_dict.keys():
                                            #Integer로 Mapping한 카테고리 dictionary 내에 값이 있는지 확인.
                                            category_value = categories_dict[element[stat_attr]]
                                            category_label = 0
                                            if category_value == CAMERA_STRING:
                                                category_label = 1
                                            elif category_value in UTILITY_ARRAY:
                                                category_label = 2
                                            elif category_value == SNS_STRING:
                                                category_label = 3
                                            elif category_value in ENTERTAINMENT_ARRAY:
                                                category_label = 4
                                            elif category_value == COMMUNICATION_STRING:
                                                category_label = 5
                                            elif category_value in GAME_ARRAY:
                                                category_label = 6
                                            elif category_value == SYSTEM_STRING:
                                                category_label = 7
                                            elif category_value in EDUCATION_ARRAY:
                                                category_label = 8
                                            elif category_value == SHOPPING_STRING:
                                                category_label = 9
                                            elif category_value in MAPS_VEHICLE_ARRAY:
                                                category_label = 10
                                            elif category_value == HEALTH_STRING:
                                                category_label = 11
                                            elif category_value == FOOD_STRING:
                                                category_label = 12
                                            elif category_value == FINANCE_STRING:
                                                category_label = 13
                                            elif category_value == BROWSER_STRING:
                                                category_label = 14
                                            sheet1.cell(idx, CELL_CATEGORY).value = category_label
                                            jsonData[user_name][str(index)]['category'] =  category_label

                                        else:
                                            #Play store에 없는 APP의 경우.
                                            uncategorizable = True
                                            non_categorizable.add(element[stat_attr])
                                        
                                    if stat_attr == 'totalTimeInForeground':
                                        sheet1.cell(idx, CELL_TOTAL_TIME_IN_FOREGROUND).value = element[stat_attr]
                                        jsonData[user_name][index]['totalTimeInForeground'] = element[stat_attr]
                                        if not uncategorizable:
                                            idx = idx + 1

<<<<<<< HEAD
                            index = index + 1
                            if idx >= MAX_CATEGORIES_BY_COROUTINE:
                                break
=======
                                index = index + 1
>>>>>>> bf13dadfa5334b93f6d4e5c2bc1edbbad207e33b
                            
            
                            
                            #case2 코루틴의 statsList 개수가 5개 미만일 경우
                            if idx - start_idx < MAX_CATEGORIES_BY_COROUTINE:
                                for i in range(0, MAX_CATEGORIES_BY_COROUTINE - (idx - start_idx)):
                                    sheet1.cell(idx, CELL_LAST_USED).value = PADDING_STATE
                                    sheet1.cell(idx, CELL_PACKAGE_NAME).value = PADDING_STATE
                                    sheet1.cell(idx, CELL_CATEGORY).value = PADDING_STATE
                                    sheet1.cell(idx, CELL_TOTAL_TIME_IN_FOREGROUND).value = PADDING_STATE
<<<<<<< HEAD
                                    jsonData[user_name][str(index)] = 0
=======

>>>>>>> bf13dadfa5334b93f6d4e5c2bc1edbbad207e33b
                                    idx += 1
    print(jsonData)
               


    wb.save(filename= 'C:\\Users\\sejin\\Documents\\GitHub\\DataManufacture\\data.xlsx')
    # print(jsonData)
    
    file_path = 'C:\\Users\\sejin\\Documents\\GitHub\\DataManufacture\\before_puls.json'
    with open(file_path, 'w') as outfile:
        json.dump(jsonData, outfile)



# print(categories_dict)